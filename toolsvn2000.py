import math
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook


def parse_number(value):
    """
    Chuyển số dạng Việt Nam hoặc dạng text sang float.
    Ví dụ:
    108,054661436777 -> 108.054661436777
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).strip()
    if not value:
        return None

    value = value.replace(",", ".")

    try:
        return float(value)
    except ValueError:
        return None


def wgs84_to_vn2000_daklak(lat, lon, h=0):
    """
    Chuyển Google Maps/GPS WGS84 sang VN2000 Đắk Lắk 108°30'.

    Đầu vào:
        lat: Vĩ độ
        lon: Kinh độ

    Đầu ra:
        X địa chính = Northing
        Y địa chính = Easting
    """

    # Ellipsoid WGS84/VN2000
    a = 6378137.0
    f = 1 / 298.257224
    e2 = 2 * f - f ** 2

    lat_rad = math.radians(lat)
    lon_rad = math.radians(lon)

    # WGS84 Geodetic -> Geocentric XYZ
    N = a / math.sqrt(1 - e2 * math.sin(lat_rad) ** 2)

    Xw = (N + h) * math.cos(lat_rad) * math.cos(lon_rad)
    Yw = (N + h) * math.cos(lat_rad) * math.sin(lon_rad)
    Zw = (N * (1 - e2) + h) * math.sin(lat_rad)

    # 7 tham số WGS84 -> VN2000 đang dùng theo bộ của anh
    dx = 191.9044143
    dy = 39.30318279
    dz = 111.4503283

    rx = math.radians(0.00928836 / 3600)
    ry = math.radians(-0.01975479 / 3600)
    rz = math.radians(0.00427372 / 3600)

    # Scale: -0.252906277 ppm
    s = -0.252906277 * 1e-6

    Xv = dx + (1 + s) * (Xw + rz * Yw - ry * Zw)
    Yv = dy + (1 + s) * (-rz * Xw + Yw + rx * Zw)
    Zv = dz + (1 + s) * (ry * Xw - rx * Yw + Zw)

    # VN2000 Geocentric XYZ -> Geodetic lat/lon
    lon_v_rad = math.atan2(Yv, Xv)
    p = math.sqrt(Xv ** 2 + Yv ** 2)

    lat_v_rad = math.atan2(Zv, p * (1 - e2))

    for _ in range(10):
        Ni = a / math.sqrt(1 - e2 * math.sin(lat_v_rad) ** 2)
        lat_v_rad = math.atan2(
            Zv + e2 * Ni * math.sin(lat_v_rad),
            p
        )

    # Chiếu Transverse Mercator / TM-3 Đắk Lắk
    central_meridian = 108.5
    k0 = 0.9999
    false_easting = 500000.0
    false_northing = 0.0

    lon0_rad = math.radians(central_meridian)
    d_lon = lon_v_rad - lon0_rad

    n = f / (2 - f)

    A = a * (1 + n ** 2 / 4 + n ** 4 / 64) / (1 + n)

    alpha = [
        0,
        1 / 2 * n - 2 / 3 * n ** 2 + 5 / 16 * n ** 3,
        13 / 48 * n ** 2 - 3 / 5 * n ** 3,
        61 / 240 * n ** 3
    ]

    beta = 2 * math.sqrt(n) / (1 + n)

    t = math.sinh(
        math.atanh(math.sin(lat_v_rad))
        - beta * math.atanh(beta * math.sin(lat_v_rad))
    )

    xi = math.atan2(t, math.cos(d_lon))

    eta = math.atanh(
        math.sin(d_lon) / math.sqrt(1 + t ** 2)
    )

    easting = false_easting + k0 * A * (
        eta + sum(
            alpha[j] * math.cos(2 * j * xi) * math.sinh(2 * j * eta)
            for j in range(1, 4)
        )
    )

    northing = false_northing + k0 * A * (
        xi + sum(
            alpha[j] * math.sin(2 * j * xi) * math.cosh(2 * j * eta)
            for j in range(1, 4)
        )
    )

    return northing, easting


def find_header_columns(ws):
    """
    Tìm cột Kinh độ, Vĩ độ trên dòng header đầu tiên.
    """
    header_row = 1
    col_lon = None
    col_lat = None

    for cell in ws[header_row]:
        value = str(cell.value).strip() if cell.value is not None else ""

        if value.lower() == "kinh độ":
            col_lon = cell.column

        if value.lower() == "vĩ độ":
            col_lat = cell.column

    return col_lat, col_lon


def process_excel(input_path):
    wb = load_workbook(input_path)
    ws = wb.active

    col_lat, col_lon = find_header_columns(ws)

    if col_lat is None or col_lon is None:
        raise ValueError("Không tìm thấy header 'Kinh độ' và 'Vĩ độ' ở dòng đầu tiên.")

    # Tạo hoặc tìm cột VN 2000
    col_vn2000 = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == "vn 2000":
            col_vn2000 = cell.column
            break

    if col_vn2000 is None:
        col_vn2000 = ws.max_column + 1
        ws.cell(row=1, column=col_vn2000).value = "VN 2000"

    count_success = 0
    count_error = 0

    for row in range(2, ws.max_row + 1):
        lat = parse_number(ws.cell(row=row, column=col_lat).value)
        lon = parse_number(ws.cell(row=row, column=col_lon).value)

        if lat is None or lon is None:
            ws.cell(row=row, column=col_vn2000).value = ""
            count_error += 1
            continue

        try:
            x_north, y_east = wgs84_to_vn2000_daklak(lat, lon)

            # Theo yêu cầu: xy= yyyyyy, xxxxxxx
            result = f"xy= {y_east:.3f}, {x_north:.3f}"

            ws.cell(row=row, column=col_vn2000).value = result
            count_success += 1

        except Exception as e:
            ws.cell(row=row, column=col_vn2000).value = f"Lỗi: {e}"
            count_error += 1

    folder = os.path.dirname(input_path)
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)

    output_path = os.path.join(folder, f"{name}_VN2000{ext}")

    wb.save(output_path)

    return output_path, count_success, count_error


def choose_file():
    file_path = filedialog.askopenfilename(
        title="Chọn file Excel",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )

    if not file_path:
        return

    entry_file.delete(0, tk.END)
    entry_file.insert(0, file_path)


def run_process():
    input_path = entry_file.get().strip()

    if not input_path:
        messagebox.showwarning("Thiếu file", "Vui lòng chọn file Excel.")
        return

    if not os.path.exists(input_path):
        messagebox.showerror("Lỗi", "File không tồn tại.")
        return

    try:
        output_path, count_success, count_error = process_excel(input_path)

        messagebox.showinfo(
            "Hoàn thành",
            f"Đã xử lý xong!\n\n"
            f"Thành công: {count_success} dòng\n"
            f"Lỗi/bỏ qua: {count_error} dòng\n\n"
            f"File kết quả:\n{output_path}"
        )

    except Exception as e:
        messagebox.showerror("Lỗi", str(e))


# =========================
# Giao diện Tkinter
# =========================

root = tk.Tk()
root.title("Chuyển WGS84 Google Maps sang VN2000 Đắk Lắk")
root.geometry("700x180")
root.resizable(False, False)

label_title = tk.Label(
    root,
    text="Chuyển Kinh độ / Vĩ độ sang VN2000 Đắk Lắk",
    font=("Arial", 14, "bold")
)
label_title.pack(pady=10)

frame_file = tk.Frame(root)
frame_file.pack(padx=15, pady=10, fill="x")

label_file = tk.Label(frame_file, text="File Excel:")
label_file.pack(side="left")

entry_file = tk.Entry(frame_file)
entry_file.pack(side="left", padx=8, fill="x", expand=True)

btn_choose = tk.Button(frame_file, text="Duyệt file", command=choose_file)
btn_choose.pack(side="left")

btn_run = tk.Button(
    root,
    text="Thực hiện chuyển VN2000",
    command=run_process,
    bg="#2e7d32",
    fg="white",
    font=("Arial", 11, "bold"),
    height=2
)
btn_run.pack(pady=15)

note = tk.Label(
    root,
    text="Yêu cầu file Excel có header: Kinh độ, Vĩ độ. Kết quả ghi dạng: xy= Y, X",
    fg="gray"
)
note.pack()

root.mainloop()